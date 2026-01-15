"""
Módulo de estilos e formatações para planilhas Excel.

Contém definições de cores, fontes e configurações de layout
utilizadas na geração das planilhas de mapeamento.
"""

from openpyxl.styles import Font, PatternFill, Alignment


class ExcelStyles:
    """Classe com definições de estilos para formatação de planilhas Excel."""
    
    # ===== CORES =====
    COLOR_HEADER_BLACK = PatternFill(
        start_color="000000", 
        end_color="000000", 
        fill_type="solid"
    )
    
    COLOR_SUBHEADER_GREEN = PatternFill(
        start_color="6aa84f", 
        end_color="6aa84f", 
        fill_type="solid"
    )
    
    COLOR_SUBHEADER_PURPLE = PatternFill(
        start_color="9900ff", 
        end_color="9900ff", 
        fill_type="solid"
    )
    
    COLOR_SUBHEADER_YELLOW = PatternFill(
        start_color="93c47d", 
        end_color="93c47d", 
        fill_type="solid"
    )
    
    COLOR_RED = PatternFill(
        start_color="FF0000", 
        end_color="FF0000", 
        fill_type="solid"
    )
    
    # ===== FONTES =====
    FONT_HEADER_WHITE = Font(color="FFFFFF", bold=True, size=11)
    FONT_HEADER_WHITE_LARGE = Font(color="FFFFFF", bold=True, size=14)
    FONT_NORMAL = Font(size=10)
    FONT_BOLD = Font(bold=True, size=10)
    
    # ===== ALINHAMENTOS =====
    ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
    ALIGN_CENTER_HORIZONTAL = Alignment(horizontal="center")
    
    # ===== LARGURAS DE COLUNAS =====
    COLUMN_WIDTHS_SUMMARY = {
        'A': 5,   # ID
        'B': 15,  # Trigger Type
        'C': 18,  # Interval frequence
        'D': 12,  # Interval days
        'E': 45,  # Movement Name
        'F': 15,  # Source System
        'G': 10,  # Source Sandbox
        'H': 12,  # Source Credentials
        'I': 15,  # Target System
        'J': 10,  # Target Sandbox
        'K': 12,  # Target Credentials
        'L': 15,  # Customization
        'M': 50,  # Notes
        'N': 8,   # No
        'O': 15,  # Email Alert
        'P': 15   # Email Every
    }
    
    COLUMN_WIDTHS_DETAIL = {
        'A': 15,
        'B': 15,
        'C': 30,
        'D': 30,
        'E': 30,
        'F': 15
    }
    
    @staticmethod
    def apply_header_style(cell, text, fill_color=None, font=None, alignment=None):
        """
        Aplica estilo a uma célula de cabeçalho.
        
        Args:
            cell: Célula do Excel
            text: Texto a ser inserido
            fill_color: Cor de fundo (PatternFill)
            font: Fonte (Font)
            alignment: Alinhamento (Alignment)
        """
        cell.value = text
        
        if fill_color:
            cell.fill = fill_color
        if font:
            cell.font = font
        if alignment:
            cell.alignment = alignment
    
    @staticmethod
    def set_column_widths(worksheet, width_dict):
        """
        Define larguras de colunas em uma planilha.
        
        Args:
            worksheet: Planilha do Excel
            width_dict: Dicionário com larguras {col_letter: width}
        """
        for col_letter, width in width_dict.items():
            worksheet.column_dimensions[col_letter].width = width

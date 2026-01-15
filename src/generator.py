"""
Módulo gerador de planilhas Excel a partir de JSONs de mapeamento Vertify.

Contém a classe MappingSpreadsheetGenerator responsável por toda a lógica
de conversão de dados JSON em planilhas Excel formatadas.
"""

import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from styles import ExcelStyles


class MappingSpreadsheetGenerator:
    """Gerador de planilha Excel a partir de JSON de mapeamentos."""
    
    def __init__(self, json_data):
        """
        Inicializa o gerador.
        
        Args:
            json_data: Dicionário com dados do JSON
        """
        self.data = json_data
        self.workbook = Workbook()
        self.styles = ExcelStyles()
        
    def create_movements_summary_tab(self):
        """Cria a aba 'Movements to migrate' com lista de todos os ObjectMaps."""
        # Remove a planilha padrão e cria nova
        if "Sheet" in self.workbook.sheetnames:
            del self.workbook["Sheet"]
        
        ws = self.workbook.create_sheet("Movements to migrate", 0)
        current_row = 1
        
        # ===== TÍTULO PRINCIPAL =====
        ws.merge_cells(f'A{current_row}:P{current_row}')
        self.styles.apply_header_style(
            ws[f'A{current_row}'],
            "Vertify movements to migrate to Digibee",
            fill_color=self.styles.COLOR_HEADER_BLACK,
            font=self.styles.FONT_HEADER_WHITE_LARGE,
            alignment=self.styles.ALIGN_CENTER
        )
        current_row += 1
        
        # ===== CUSTOMER INFO =====
        ws['A2'] = "Customer:"
        ws['B2'] = "?"
        ws['A3'] = "Key Documents:"
        ws['B3'] = "?"
        current_row = 4
        
        # ===== HEADER VERDE - VERTIFY =====
        ws.merge_cells(f'A{current_row}:P{current_row}')
        self.styles.apply_header_style(
            ws[f'A{current_row}'],
            "Vertify",
            fill_color=self.styles.COLOR_SUBHEADER_GREEN,
            font=self.styles.FONT_BOLD,
            alignment=self.styles.ALIGN_CENTER
        )
        current_row += 1
        
        # ===== SUB-HEADERS - SOURCE E TARGET SYSTEM =====
        ws.merge_cells(f'F{current_row}:H{current_row}')
        self.styles.apply_header_style(
            ws[f'F{current_row}'],
            "Source System",
            fill_color=self.styles.COLOR_SUBHEADER_GREEN,
            alignment=self.styles.ALIGN_CENTER
        )
        
        ws.merge_cells(f'I{current_row}:L{current_row}')
        self.styles.apply_header_style(
            ws[f'I{current_row}'],
            "Target System",
            fill_color=self.styles.COLOR_SUBHEADER_GREEN,
            alignment=self.styles.ALIGN_CENTER
        )
        current_row += 1
        
        # ===== HEADERS DAS COLUNAS =====
        headers = [
            "ID", "Trigger Type", "Interval frequence", "Interval days", "Movement Name",
            "System", "Sandbox", "Credentials", "System", "Sandbox", "Credentials",
            "Customization", "Notes", "No", "Email Alert", "Email Every"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num)
            self.styles.apply_header_style(
                cell,
                header,
                fill_color=self.styles.COLOR_SUBHEADER_GREEN,
                font=self.styles.FONT_BOLD,
                alignment=self.styles.ALIGN_CENTER
            )
        current_row += 1
        
        # ===== ADICIONAR DADOS DOS OBJECTMAPS =====
        objects_map = self.data.get("ObjectsMap", [])
        for idx, obj_map in enumerate(objects_map, 1):
            source_system = obj_map.get("SourceSystemName", "N/A")
            target_system = obj_map.get("TargetSystemName", "N/A")
            trigger_type = "Collect & Move? / Collect?"
            
            row_data = [
                idx, trigger_type, "at 00:00 AM", "every ?", obj_map.get("Name", "N/A"),
                source_system, "TRUE/FALSE", "TRUE/FALSE", target_system, "TRUE/FALSE",
                "TRUE/FALSE", "TRUE/FALSE", "", "", "", ""
            ]
            
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=current_row, column=col_num).value = value
            
            current_row += 1
        
        # ===== AJUSTAR LARGURA DAS COLUNAS =====
        self.styles.set_column_widths(ws, self.styles.COLUMN_WIDTHS_SUMMARY)
        
        # Congelar linhas de cabeçalho
        ws.freeze_panes = "A7"
        
    def create_object_map_tab(self, idx, obj_map):
        """
        Cria uma aba detalhada para um ObjectMap específico.
        
        Args:
            idx: Índice do ObjectMap
            obj_map: Dicionário com dados do ObjectMap
        """
        # Nome da aba (limitado a 31 caracteres do Excel e sem caracteres inválidos)
        name = obj_map.get('Name', 'Unknown')
        name = self._sanitize_sheet_name(name)
        tab_name = f"{idx} - {name}"
        if len(tab_name) > 31:
            tab_name = f"{idx} - {name[:22]}..."
        
        ws = self.workbook.create_sheet(tab_name)
        
        current_row = 1
        
        # ===== SEÇÕES =====
        current_row = self._add_api_request_section(ws, obj_map, current_row)
        current_row += 2
        
        current_row = self._add_merge_section(ws, obj_map, current_row)
        current_row += 2
        
        current_row = self._add_filter_section(ws, obj_map, current_row)
        current_row += 2
        
        current_row = self._add_field_mapping_section(ws, obj_map, current_row)
        
        # Ajustar larguras
        self.styles.set_column_widths(ws, self.styles.COLUMN_WIDTHS_DETAIL)
        
    def _sanitize_sheet_name(self, name):
        """
        Remove caracteres inválidos de nomes de abas do Excel.
        
        Args:
            name: Nome original
            
        Returns:
            Nome sanitizado
        """
        replacements = {
            ":": " ",
            ">>": "to",
            "\\": " ",
            "/": " ",
            "?": "",
            "*": "",
            "[": "(",
            "]": ")"
        }
        
        for old, new in replacements.items():
            name = name.replace(old, new)
        
        return name
        
    def _add_api_request_section(self, ws, obj_map, start_row):
        """Adiciona seção de API Request."""
        ws.merge_cells(f'A{start_row}:F{start_row}')
        self.styles.apply_header_style(
            ws[f'A{start_row}'],
            "API Request",
            fill_color=self.styles.COLOR_RED,
            font=self.styles.FONT_HEADER_WHITE,
            alignment=self.styles.ALIGN_CENTER
        )
        start_row += 1
        
        ws.merge_cells(f'A{start_row}:F{start_row}')
        self.styles.apply_header_style(
            ws[f'A{start_row}'],
            "MIGRATION SYTEM: VERTIFY",
            fill_color=self.styles.COLOR_SUBHEADER_GREEN,
            alignment=self.styles.ALIGN_CENTER
        )
        start_row += 1
        
        headers = ["system", "type", "path/connection string", "request example", "response example", "notes"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_num)
            self.styles.apply_header_style(
                cell,
                header,
                fill_color=self.styles.COLOR_SUBHEADER_GREEN
            )
        start_row += 1
        
        source_system = obj_map.get("SourceSystemName", "")
        target_system = obj_map.get("TargetSystemName", "")
        
        ws.cell(row=start_row, column=1).value = source_system
        ws.cell(row=start_row, column=2).value = "REST"
        start_row += 1
        
        ws.cell(row=start_row, column=1).value = target_system
        ws.cell(row=start_row, column=2).value = "REST"
        
        return start_row + 1
        
    def _add_merge_section(self, ws, obj_map, start_row):
        """Adiciona seção de Merge."""
        ws.merge_cells(f'A{start_row}:F{start_row}')
        self.styles.apply_header_style(
            ws[f'A{start_row}'],
            "Merge",
            fill_color=self.styles.COLOR_SUBHEADER_GREEN,
            font=self.styles.FONT_HEADER_WHITE,
            alignment=self.styles.ALIGN_CENTER
        )
        start_row += 1
        
        ws.merge_cells(f'A{start_row}:C{start_row}')
        self.styles.apply_header_style(
            ws[f'A{start_row}'],
            "Vertify",
            fill_color=self.styles.COLOR_SUBHEADER_GREEN,
            alignment=self.styles.ALIGN_CENTER_HORIZONTAL
        )
        start_row += 1
        
        ws.cell(row=start_row, column=1).value = "rules"
        ws.cell(row=start_row, column=1).fill = self.styles.COLOR_SUBHEADER_GREEN
        
        ws.merge_cells(f'D{start_row-1}:F{start_row-1}')
        self.styles.apply_header_style(
            ws[f'D{start_row-1}'],
            "Digibee",
            fill_color=self.styles.COLOR_SUBHEADER_PURPLE,
            alignment=self.styles.ALIGN_CENTER_HORIZONTAL
        )
        
        ws.cell(row=start_row, column=4).value = "rules"
        ws.cell(row=start_row, column=4).fill = self.styles.COLOR_SUBHEADER_PURPLE
        start_row += 1
        
        merge_record = obj_map.get("MergeRecord", False)
        merge_fields = obj_map.get("ObjectsMapMergeField", [])
        
        if merge_record and merge_fields:
            merge_info = []
            for field in merge_fields:
                merge_info.append(
                    f"{field.get('MergeField', '')}: "
                    f"{field.get('SourcePropertyName', '')} -> "
                    f"{field.get('TargetPropertyName', '')}"
                )
            ws.cell(row=start_row, column=1).value = "\n".join(merge_info)
        else:
            ws.cell(row=start_row, column=1).value = "No merge"
        
        ws.cell(row=start_row, column=4).value = "N/A"
        
        return start_row + 1
        
    def _add_filter_section(self, ws, obj_map, start_row):
        """Adiciona seção de Filter."""
        ws.merge_cells(f'A{start_row}:F{start_row}')
        self.styles.apply_header_style(
            ws[f'A{start_row}'],
            "Filter",
            fill_color=self.styles.COLOR_HEADER_BLACK,
            font=self.styles.FONT_HEADER_WHITE,
            alignment=self.styles.ALIGN_CENTER
        )
        start_row += 1
        
        ws.merge_cells(f'A{start_row}:C{start_row}')
        self.styles.apply_header_style(
            ws[f'A{start_row}'],
            "Vertify",
            fill_color=self.styles.COLOR_SUBHEADER_GREEN,
            alignment=self.styles.ALIGN_CENTER_HORIZONTAL
        )
        
        ws.merge_cells(f'D{start_row}:F{start_row}')
        self.styles.apply_header_style(
            ws[f'D{start_row}'],
            "Digibee",
            fill_color=self.styles.COLOR_SUBHEADER_PURPLE,
            alignment=self.styles.ALIGN_CENTER_HORIZONTAL
        )
        start_row += 1
        
        ws.cell(row=start_row, column=1).value = "FILTER"
        ws.cell(row=start_row, column=1).fill = self.styles.COLOR_SUBHEADER_GREEN
        
        ws.merge_cells(f'D{start_row}:F{start_row}')
        cell = ws[f'D{start_row}']
        cell.value = ""
        cell.fill = self.styles.COLOR_SUBHEADER_PURPLE
        cell.alignment = self.styles.ALIGN_CENTER_HORIZONTAL
        start_row += 1
        
        filter_headers_left = ["path.field", "condition", "value"]
        filter_headers_right = ["path/table/alias", "field", "query relation"]
        
        for col_num, header in enumerate(filter_headers_left, 1):
            self.styles.apply_header_style(
                ws.cell(row=start_row, column=col_num),
                header,
                fill_color=self.styles.COLOR_SUBHEADER_GREEN
            )
            
        for col_num, header in enumerate(filter_headers_right, 4):
            self.styles.apply_header_style(
                ws.cell(row=start_row, column=col_num),
                header,
                fill_color=self.styles.COLOR_SUBHEADER_PURPLE
            )
        start_row += 1
        
        filters = obj_map.get("ObjectsMapFilter", [])
        if filters:
            for filter_item in filters:
                ws.cell(row=start_row, column=1).value = filter_item.get("SourcePropertyName", "")
                ws.cell(row=start_row, column=2).value = filter_item.get("FilterOperator", "")
                ws.cell(row=start_row, column=3).value = filter_item.get("Value", "")
                start_row += 1
        else:
            ws.cell(row=start_row, column=1).value = "No filter"
            start_row += 1
        
        return start_row
        
    def _add_field_mapping_section(self, ws, obj_map, start_row):
        """Adiciona seção de Field Mapping."""
        ws.merge_cells(f'A{start_row}:F{start_row}')
        self.styles.apply_header_style(
            ws[f'A{start_row}'],
            "Field Mapping",
            fill_color=self.styles.COLOR_HEADER_BLACK,
            font=self.styles.FONT_HEADER_WHITE,
            alignment=self.styles.ALIGN_CENTER
        )
        start_row += 1
        
        ws.merge_cells(f'A{start_row}:F{start_row}')
        self.styles.apply_header_style(
            ws[f'A{start_row}'],
            "MIGRATION SYTEM: VERTIFY",
            fill_color=self.styles.COLOR_SUBHEADER_GREEN,
            alignment=self.styles.ALIGN_CENTER_HORIZONTAL
        )
        start_row += 1
        
        headers = ["move", "type", "details", "source path.field", "target path.field", "notes"]
        for col_num, header in enumerate(headers, 1):
            self.styles.apply_header_style(
                ws.cell(row=start_row, column=col_num),
                header,
                fill_color=self.styles.COLOR_SUBHEADER_GREEN
            )
        start_row += 1
        
        properties = obj_map.get("PropertiesMap", [])
        for prop in properties:
            move_action = prop.get("MoveAction", "") or "OnAddUpdate"
            prop_type = prop.get("Type", "Map")
            target_prop = prop.get("TargetPropertyName", "")
            
            transformations = prop.get("PropertiesMapTransformation", [])
            source_prop = ""
            details = ""
            
            if transformations:
                first_transform = transformations[0]
                source_prop = first_transform.get("SourcePropertyName", "")
                rule_type = first_transform.get("RuleType", "")
                
                details = self._get_transformation_details(first_transform, rule_type)
            
            ws.cell(row=start_row, column=1).value = move_action
            ws.cell(row=start_row, column=2).value = prop_type
            ws.cell(row=start_row, column=3).value = details
            ws.cell(row=start_row, column=4).value = source_prop
            ws.cell(row=start_row, column=5).value = target_prop
            start_row += 1
        
        return start_row
    
    def _get_transformation_details(self, transform, rule_type):
        """
        Extrai detalhes da transformação baseado no tipo de regra.
        
        Args:
            transform: Dicionário com dados da transformação
            rule_type: Tipo da regra
            
        Returns:
            String com detalhes formatados
        """
        if rule_type == "Value":
            return f"Value: {transform.get('Value', '')}"
        elif rule_type == "Convert":
            return f"Convert List: {transform.get('ProjectConvertListName', '')}"
        elif rule_type == "Condition":
            return "Conditional Logic"
        elif rule_type == "Date":
            return f"Date Format: {transform.get('DateFormat', '')}"
        return ""
        
    def generate_to_bytes(self):
        """
        Gera a planilha e retorna como bytes.
        
        Returns:
            bytes: Conteúdo da planilha Excel
        """
        # Criar aba de resumo
        self.create_movements_summary_tab()
        
        # Criar abas para cada ObjectMap
        objects_map = self.data.get("ObjectsMap", [])
        for idx, obj_map in enumerate(objects_map, 1):
            self.create_object_map_tab(idx, obj_map)
        
        # Salvar em BytesIO
        output = io.BytesIO()
        self.workbook.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def get_statistics(self):
        """
        Retorna estatísticas sobre o JSON processado.
        
        Returns:
            dict: Dicionário com estatísticas
        """
        objects_map = self.data.get("ObjectsMap", [])
        
        total_properties = sum(
            len(obj.get("PropertiesMap", [])) 
            for obj in objects_map
        )
        
        total_filters = sum(
            len(obj.get("ObjectsMapFilter", [])) 
            for obj in objects_map
        )
        
        return {
            "total_objectmaps": len(objects_map),
            "total_properties": total_properties,
            "total_filters": total_filters
        }
